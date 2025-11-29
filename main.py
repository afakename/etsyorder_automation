# main.py - Complete with open orders filtering by status field
import sys
import argparse
import json
from datetime import datetime
from pathlib import Path
from config import Config
from logger import setup_logger
import pandas as pd

# Import the existing etsy_api_connector
from etsy_api_connector import EtsyAPIConnector
from filename_generator import FilenameGenerator
from file_database import FileDatabase

class EtsyAutomation:
    # SKU categories
    PEACE_LOVE_HOPE_SKUS = {
        'FckFlkEarring-RR-01', 'FuckFlkOrn-Mirr-03', 'FuckFlkOrn-Mirr-04',
        'FuckFlkOrn-Mirr-05', 'FuckFlkOrn-RR-01', 'FuckFlkOrn-RR-02',
        'HopeFlkOrn-Mirr-01', 'HopeFlkOrn-RR-01', 'LoveFlkOrn-Mirr-01',
        'LoveFlkOrn-RR-01', 'PeaceFlkOrn-Mirr-01', 'PeaceFlkOrn-RR-01',
        'PeacLovHopeFlkOrn-Mirr-02', 'PeacLovHopeFlkOrn-RR-01', 'SnoFlkEarring-01'
    }
    
    OTHER_ORDERS_SKUS = {
        'CandyHeartEarring-Dangle-02', 'CandyHeartEarring-Studs-01',
        'MMEarHld-01', 'PersPhotoMag-01', 'PersPhotoMag-02',
        'ThermPrntStand-3D-01', 'USMC-AAV-01', 'USMC-AAVPen-02'
    }
    
    def __init__(self, mode="open_only", days_back=None):
        # Set up logging
        Config.get_log_path().mkdir(parents=True, exist_ok=True)
        
        # Create timestamped run folder
        self.run_timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        self.output_folder = Config.get_output_path() / self.run_timestamp
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        log_file = self.output_folder / f"automation.log"
        
        self.logger = setup_logger("main", log_file)
        self.logger.info("=" * 50)
        self.logger.info(f"Starting Etsy Automation - Run: {self.run_timestamp}")
        self.logger.info(f"Mode: {mode}")
        if days_back:
            self.logger.info(f"Days back: {days_back}")
        self.logger.info("=" * 50)
        
        # Store mode settings
        self.mode = mode
        self.days_back = days_back
        
        # Initialize components
        self.etsy_client = EtsyAPIConnector()
        self.filename_generator = FilenameGenerator(self.logger)
        self.file_database = FileDatabase(self.logger)
        
        # Load previous run data for status tracking
        self.previous_orders = self.load_previous_orders()
    
    def load_previous_orders(self):
        """Load order IDs from previous run"""
        tracking_file = Config.get_output_path() / "last_run_orders.json"
        
        if tracking_file.exists():
            try:
                with open(tracking_file, 'r') as f:
                    data = json.load(f)
                    self.logger.info(f"Loaded {len(data.get('order_ids', []))} order IDs from previous run")
                    return set(data.get('order_ids', []))
            except Exception as e:
                self.logger.warning(f"Could not load previous orders: {e}")
                return set()
        else:
            self.logger.info("No previous run found - all orders will be marked as 'New'")
            return set()
    
    def save_current_orders(self, order_ids):
        """Save current order IDs for next run"""
        tracking_file = Config.get_output_path() / "last_run_orders.json"
        
        try:
            with open(tracking_file, 'w') as f:
                json.dump({
                    'timestamp': self.run_timestamp,
                    'order_ids': list(order_ids)
                }, f, indent=2)
            
            # Also save a copy in the run folder
            run_tracking_file = self.output_folder / "last_run_orders.json"
            with open(run_tracking_file, 'w') as f:
                json.dump({
                    'timestamp': self.run_timestamp,
                    'order_ids': list(order_ids)
                }, f, indent=2)
                
            self.logger.info(f"Saved {len(order_ids)} order IDs for next run")
        except Exception as e:
            self.logger.error(f"Could not save order tracking: {e}")
    
    def normalize_filename(self, filename):
        """Normalize filename for comparison - handles case variations and spelling variations"""
        # Convert to lowercase and normalize common variations
        normalized = filename.lower()
        # Standardize spacing
        normalized = ' '.join(normalized.split())
        
        # Standardize flake/flk variations to 'flk'
        normalized = normalized.replace('flake', 'flk')
        
        # Additional normalizations if needed in the future
        # normalized = normalized.replace('star', 'str') # example
        
        return normalized
    
    def run(self):
        """Main execution flow"""
        try:
            # Get all orders based on mode
            all_orders = self.get_all_orders()
            
            if not all_orders:
                self.logger.info("No orders found")
                print("\nNo orders found.")
                return
            
            # Categorize orders
            categorized = self.categorize_orders(all_orders)
            
            # Process workflow orders (MS/RR)
            workflow_results = self.process_workflow_orders(categorized['workflow'])
            
            # Process other order types
            peace_love_hope_results = self.process_other_orders(categorized['peace_love_hope'])
            other_orders_results = self.process_other_orders(categorized['other'])
            
            # Generate reports
            self.generate_reports(workflow_results, peace_love_hope_results, other_orders_results)
            
            # Save order IDs for next run
            all_order_ids = set()
            for order in all_orders:
                all_order_ids.add(str(order.get('receipt_id', '')))
            self.save_current_orders(all_order_ids)
            
            self.logger.info("Automation completed successfully")
            print(f"\nAutomation completed successfully!")
            print(f"Output folder: {self.output_folder}")
            
        except Exception as e:
            self.logger.error(f"Automation failed: {e}")
            raise
    
    def get_all_orders(self):
        """Get orders based on mode setting"""
        if self.mode == "open_only":
            # Open orders mode - filters by status != Complete
            days = self.days_back or 90
            self.logger.info(f"Fetching open orders only (last {days} days)")
            all_orders = self.etsy_client.get_open_orders(days_back=days, limit=500)
        else:
            # Default: time_based mode
            days = self.days_back or 30
            self.logger.info(f"Fetching orders from last {days} days (all statuses)")
            all_orders = self.etsy_client.get_recent_orders(days_back=days, limit=500)
        
        if not all_orders:
            return []
        
        self.logger.info(f"Found {len(all_orders)} total orders")
        return all_orders
    
    def categorize_orders(self, orders):
        """Categorize orders by SKU type"""
        workflow = []
        peace_love_hope = []
        other = []
        
        for order in orders:
            order_workflow = []
            order_peace = []
            order_other = []
            
            if 'transactions' in order:
                for transaction in order['transactions']:
                    sku = transaction.get('sku', '')
                    
                    if sku in Config.WORKFLOW_SKUS:
                        order_workflow.append(transaction)
                    elif sku in self.PEACE_LOVE_HOPE_SKUS:
                        order_peace.append(transaction)
                    elif sku in self.OTHER_ORDERS_SKUS:
                        order_other.append(transaction)
            
            # Add to appropriate category if items found
            if order_workflow:
                workflow_order = order.copy()
                workflow_order['transactions'] = order_workflow
                workflow.append(workflow_order)
            
            if order_peace:
                peace_order = order.copy()
                peace_order['transactions'] = order_peace
                peace_love_hope.append(peace_order)
            
            if order_other:
                other_order = order.copy()
                other_order['transactions'] = order_other
                other.append(other_order)
        
        self.logger.info(f"Categorized: {len(workflow)} workflow, {len(peace_love_hope)} peace/love/hope, {len(other)} other")
        
        return {
            'workflow': workflow,
            'peace_love_hope': peace_love_hope,
            'other': other
        }
    
    def check_file_status(self, filename):
        """
        Determine if design needs to be made, updated, or already exists.
        VERSION-AWARE LOGIC: Finds all versions and prioritizes highest version number.
        Returns: (status, file_path, update_details)
        """
        # Step 1: Check for EXACT match (same name, same version, same everything)
        normalized_search = self.normalize_filename(filename)

        for indexed_filename, file_path in self.file_database.file_index.items():
            if self.normalize_filename(indexed_filename) == normalized_search:
                self.logger.info(f"EXACT MATCH found: {indexed_filename}")
                return 'exists', file_path, None

        # Step 2: Use fuzzy search to find all similar designs (ignores version numbers)
        # This will find "Amber Star", "Amber 2 Star", "Amber 3 Star" when searching for "Amber Star"
        # and return them sorted by version (highest first)
        fuzzy_matches = self.file_database.fuzzy_search(filename)

        if fuzzy_matches:
            # Get the highest version match (first in sorted list)
            best_match = fuzzy_matches[0]
            best_match_filename = best_match.stem

            self.logger.info(f"FUZZY MATCH found: {best_match_filename} (highest version)")

            # Parse the ordered filename to compare details
            ordered_center = self.extract_center_from_filename(filename)
            ordered_year = self.extract_year(filename)

            # Parse the matched filename to compare details
            matched_center = self.extract_center_from_filename(best_match_filename)
            matched_year = self.extract_year(best_match_filename)

            # If center or year differs, it's an update
            if ordered_center != matched_center or ordered_year != matched_year:
                changes = []
                if matched_center != ordered_center:
                    changes.append(f"Center: {matched_center.capitalize()} → {ordered_center.capitalize()}")
                if matched_year != ordered_year:
                    old_year = matched_year if matched_year else "No Year"
                    new_year = ordered_year if ordered_year else "No Year"
                    changes.append(f"Year: {old_year} → {new_year}")

                update_details = " | ".join(changes)
                self.logger.info(f"UPDATE needed: {best_match_filename} → {filename} ({update_details})")
                return 'update', best_match, update_details
            else:
                # Same center and year - this is essentially the same design
                # Return 'exists' with the highest version file
                self.logger.info(f"MATCH found (version {self.file_database.get_version_number(best_match_filename)}): {best_match_filename}")
                return 'exists', best_match, None

        # Step 3: No matches found - needs to be made
        self.logger.info(f"NO MATCH found for: {filename} - needs to be MADE")
        return 'make', None, None
    
    def extract_center_from_filename(self, filename):
        """Extract center design from filename (handles case variations)"""
        filename_lower = filename.lower()
        
        # Check for star variations
        if 'star' in filename_lower:
            return 'star'
        
        # Check for flake/flk variations
        if 'flake' in filename_lower or 'flk' in filename_lower:
            return 'flk'
        
        # Check for year (means star center for RR)
        parts = filename.split()
        for part in parts:
            if part.isdigit() and len(part) == 4:
                return 'star'
        
        return 'star'  # Default
    
    def extract_center(self, filename, variations):
        """Extract center piece from filename and variations for output"""
        filename_lower = filename.lower()
        
        if 'star' in filename_lower:
            return 'Star'
        elif 'flake' in filename_lower or 'flk' in filename_lower:
            return 'Flk'
        
        # For RR variants, check if it's a year
        parts = filename.split()
        for part in parts:
            if part.isdigit() and len(part) == 4:
                return part  # It's a year
        
        return 'star'  # Default for RR
    
    def extract_year(self, filename):
        """Extract year from filename"""
        parts = filename.split()
        for part in parts:
            if part.isdigit() and len(part) == 4:
                return part
        return ''
    
    def get_order_status(self, order_id):
        """Determine if order is new or previously pulled"""
        return "New" if str(order_id) not in self.previous_orders else "Previously Pulled"
    
    def process_workflow_orders(self, orders):
        """Process workflow orders (MS/RR) with file status checking"""
        all_orders = []
        needs_made = []
        needs_updated = []
        file_locations = []
        
        for order in orders:
            customer_name = order.get('name', 'Unknown Customer')
            order_id = order.get('receipt_id', 'Unknown')
            message = order.get('message_from_buyer', '')
            
            # Determine order status
            order_status = self.get_order_status(order_id)
            
            for transaction in order.get('transactions', []):
                # Generate filename
                filename = self.filename_generator.generate_filename(transaction)
                
                if not filename:
                    continue
                
                # Check file status with UPDATED logic
                status, file_path, update_details = self.check_file_status(filename)
                
                # Extract variations first
                variations = self.filename_generator.extract_variations(transaction.get('variations', []))
                
                # Check for preview request in VARIATIONS (not message)
                # Check multiple possible field names
                preview = '0'
                preview_requested = False
                
                # Log all variations for debugging
                self.logger.debug(f"Order {order_id} variations: {variations}")
                
                # Check multiple possible variation names
                for key, value in variations.items():
                    if key and value:
                        key_lower = key.lower()
                        value_lower = value.lower()
                        # Check if this is a preview-related field
                        if 'preview' in key_lower or 'proof' in key_lower or 'mock' in key_lower:
                            self.logger.info(f"Found preview variation: {key} = {value}")
                            if 'yes' in value_lower:
                                preview = '1'
                                preview_requested = True
                                self.logger.info(f"Preview requested for order {order_id}")
                                break
                
                center = self.extract_center(filename, variations)
                year = self.extract_year(filename)
                
                order_data = {
                    'order_status': order_status,
                    'order_id': order_id,
                    'customer_name': customer_name,
                    'sku': transaction.get('sku', ''),
                    'file_status': status,
                    'file_path': str(file_path) if file_path else 'NOT FOUND',
                    'update_details': update_details if update_details else '',
                    'quantity': transaction.get('quantity', 1),
                    'price': self.format_price(transaction.get('price', {})),
                    'personalization': variations.get('Personalization', ''),
                    'center': center,
                    'year': year if year else 'No',
                    'preview': preview,
                    'preview_requested': preview_requested,
                    'message': message[:100] if message else '',
                    'generated_filename': filename
                }
                
                all_orders.append(order_data)
                
                if status == 'make':
                    needs_made.append(order_data)
                elif status == 'update':
                    needs_updated.append(order_data)
                else:
                    file_locations.append(order_data)
        
        return {
            'all_orders': all_orders,
            'needs_made': needs_made,
            'needs_updated': needs_updated,
            'file_locations': file_locations
        }
    
    def process_other_orders(self, orders):
        """Process peace/love/hope and other orders"""
        processed = []
        
        for order in orders:
            customer_name = order.get('name', 'Unknown Customer')
            order_id = order.get('receipt_id', 'Unknown')
            message = order.get('message_from_buyer', '')
            
            # Determine order status
            order_status = self.get_order_status(order_id)
            
            for transaction in order.get('transactions', []):
                # Extract variation details
                variations = self.filename_generator.extract_variations(transaction.get('variations', []))
                variation_text = ', '.join([f"{k}: {v}" for k, v in variations.items()])
                
                order_data = {
                    'order_status': order_status,
                    'order_id': order_id,
                    'customer_name': customer_name,
                    'product_title': transaction.get('title', 'Unknown Product'),
                    'sku': transaction.get('sku', ''),
                    'quantity': transaction.get('quantity', 1),
                    'price': self.format_price(transaction.get('price', {})),
                    'message': message[:200] if message else '',
                    'variation_selected': variation_text,
                }
                
                processed.append(order_data)
        
        return processed
    
    def get_days_since_modified(self, file_path):
        """Calculate days since file was last modified"""
        try:
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                return 'N/A'

            modified_time = datetime.fromtimestamp(file_path_obj.stat().st_mtime)
            days_ago = (datetime.now() - modified_time).days
            return days_ago
        except Exception as e:
            self.logger.warning(f"Could not get modified time for {file_path}: {e}")
            return 'N/A'

    def format_price(self, price_data):
        """Format price from Etsy API format"""
        if isinstance(price_data, dict):
            amount = price_data.get('amount', 0)
            divisor = price_data.get('divisor', 100)
            return f"${amount / divisor:.2f}"
        return str(price_data)
    
    def apply_customer_grouping(self, worksheet, data, customer_col_name):
        """
        Apply alternating row colors by customer to visually group orders
        
        Args:
            worksheet: openpyxl worksheet object
            data: DataFrame containing the data
            customer_col_name: Name of the customer column
        """
        from openpyxl.styles import PatternFill
        
        # Define two alternating colors
        color1 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
        color2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        
        if customer_col_name not in data.columns:
            return
        
        # Start with first color
        current_color = color1
        last_customer = None
        
        # Apply colors (start at row 2 because row 1 is headers)
        for idx in range(len(data)):
            row_num = idx + 2  # Excel rows start at 1, headers are row 1
            customer = data.iloc[idx][customer_col_name]
            
            # Switch color when customer changes
            if last_customer is not None and customer != last_customer:
                current_color = color1 if current_color == color2 else color2
            
            # Apply color to entire row
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col).fill = current_color
            
            last_customer = customer
    
    def generate_reports(self, workflow_results, peace_love_hope_results, other_orders_results):
        """Generate Excel reports with multiple sheets"""
        output_file = self.output_folder / "etsy_orders.xlsx"
        
        # Separate MS and RR items, then separate Make vs Update
        ms_make = [item for item in workflow_results['needs_made'] if 'MS' in item['sku']]
        ms_update = [item for item in workflow_results['needs_updated'] if 'MS' in item['sku']]
        rr_make = [item for item in workflow_results['needs_made'] if 'MS' not in item['sku']]
        rr_update = [item for item in workflow_results['needs_updated'] if 'MS' not in item['sku']]
        
        # Collect all preview-requested items for Preview sheet
        preview_items = []
        for item in workflow_results['all_orders']:
            if item.get('preview_requested', False):
                self.logger.info(f"Adding to preview sheet: {item['generated_filename']}")
                # Determine production location based on file status and SKU
                if item['file_status'] == 'exists':
                    location = 'Already Made'
                    file_path_display = item['file_path']
                elif item['file_status'] == 'update':
                    if 'MS' in item['sku']:
                        location = 'MS - Needs Updated'
                    else:
                        location = 'RR - Needs Updated'
                    file_path_display = ''
                else:  # make
                    if 'MS' in item['sku']:
                        location = 'MS - Needs Made'
                    else:
                        location = 'RR - Needs Made'
                    file_path_display = ''
                
                preview_items.append({
                    'Status': item['order_status'],
                    'Order ID': item['order_id'],
                    'Customer': item['customer_name'],
                    'Name': item['personalization'],
                    'SKU': item['sku'],
                    'Production Location': location,
                    'File Path': file_path_display,
                    'Quantity': item['quantity'],
                    'Generated Filename': item['generated_filename'],
                    'Message': item['message']
                })
        
        self.logger.info(f"Total preview requests found: {len(preview_items)}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Preview Requests sheet - FIRST SHEET (always create, even if empty)
            if preview_items:
                df_preview = pd.DataFrame(preview_items)
            else:
                # Create empty dataframe with column headers
                df_preview = pd.DataFrame(columns=[
                    'Status', 'Order ID', 'Customer', 'Name', 'SKU', 
                    'Production Location', 'File Path', 'Quantity', 
                    'Generated Filename', 'Message'
                ])
            
            df_preview.to_excel(writer, sheet_name='Preview Requests', index=False)
            
            # Apply customer grouping colors
            if len(preview_items) > 0:
                worksheet = writer.sheets['Preview Requests']
                self.apply_customer_grouping(worksheet, df_preview, 'Customer')
            
            # Add hyperlinks to file paths for Already Made items
            if len(preview_items) > 0:
                worksheet = writer.sheets['Preview Requests']
                file_path_col = df_preview.columns.get_loc('File Path') + 1
                for idx in range(len(df_preview)):
                    file_path = df_preview.iloc[idx]['File Path']
                    if file_path and file_path != 'NOT FOUND' and file_path != '':
                        cell = worksheet.cell(row=idx+2, column=file_path_col)
                        cell.hyperlink = file_path
                        cell.style = 'Hyperlink'
            
            # MS Needs Made sheet
            if ms_make:
                ms_make_data = []
                for item in ms_make:
                    ms_make_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Name': item['personalization'],
                        'Center': item['center'],
                        'Year': item['year'],
                        'Preview': item['preview'],
                        'Quantity': item['quantity'],
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'SKU': item['sku'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Generated Filename': item['generated_filename']
                    })
                df_ms_make = pd.DataFrame(ms_make_data)
                df_ms_make.to_excel(writer, sheet_name='MS - Needs Made', index=False)
                
                # Apply customer grouping colors
                worksheet = writer.sheets['MS - Needs Made']
                self.apply_customer_grouping(worksheet, df_ms_make, 'Customer')
            
            # MS Needs Updated sheet
            if ms_update:
                ms_update_data = []
                for item in ms_update:
                    ms_update_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Name': item['personalization'],
                        'Center': item['center'],
                        'Year': item['year'],
                        'Preview': item['preview'],
                        'Quantity': item['quantity'],
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'SKU': item['sku'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Generated Filename': item['generated_filename'],
                        'Update Details': item.get('update_details', '')
                    })
                df_ms_update = pd.DataFrame(ms_update_data)
                df_ms_update.to_excel(writer, sheet_name='MS - Needs Updated', index=False)
                
                # Apply customer grouping colors
                worksheet = writer.sheets['MS - Needs Updated']
                self.apply_customer_grouping(worksheet, df_ms_update, 'Customer')
            
            # RR Needs Made sheet
            if rr_make:
                rr_make_data = []
                for item in rr_make:
                    rr_make_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Name': item['personalization'],
                        'Center': item['center'],
                        'Preview': item['preview'],
                        'Quantity': item['quantity'],
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'SKU': item['sku'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Generated Filename': item['generated_filename']
                    })
                df_rr_make = pd.DataFrame(rr_make_data)
                df_rr_make.to_excel(writer, sheet_name='RR - Needs Made', index=False)
                
                # Apply customer grouping colors
                worksheet = writer.sheets['RR - Needs Made']
                self.apply_customer_grouping(worksheet, df_rr_make, 'Customer')
            
            # RR Needs Updated sheet
            if rr_update:
                rr_update_data = []
                for item in rr_update:
                    rr_update_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Name': item['personalization'],
                        'Center': item['center'],
                        'Preview': item['preview'],
                        'Quantity': item['quantity'],
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'SKU': item['sku'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Generated Filename': item['generated_filename'],
                        'Update Details': item.get('update_details', '')
                    })
                df_rr_update = pd.DataFrame(rr_update_data)
                df_rr_update.to_excel(writer, sheet_name='RR - Needs Updated', index=False)
                
                # Apply customer grouping colors
                worksheet = writer.sheets['RR - Needs Updated']
                self.apply_customer_grouping(worksheet, df_rr_update, 'Customer')
            
            # Already Made sheet with hyperlinked file paths
            if workflow_results['file_locations']:
                made_data = []
                for item in workflow_results['file_locations']:
                    days_since_modified = self.get_days_since_modified(item['file_path'])
                    made_data.append({
                        'Status': item['order_status'],
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'Name': item['personalization'],
                        'SKU': item['sku'],
                        'Generated Filename': item['generated_filename'],
                        'File Path': item['file_path'],
                        'Days Since Modified': days_since_modified,
                        'Quantity': item['quantity'],
                        'Preview': 'Yes' if item.get('preview_requested', False) else 'No'
                    })
                df_made = pd.DataFrame(made_data)
                df_made.to_excel(writer, sheet_name='Already Made', index=False)
                
                # Apply customer grouping colors
                worksheet = writer.sheets['Already Made']
                self.apply_customer_grouping(worksheet, df_made, 'Customer')
                
                # Add hyperlinks to file paths
                file_path_col = df_made.columns.get_loc('File Path') + 1
                for idx in range(len(df_made)):
                    file_path = df_made.iloc[idx]['File Path']
                    if file_path != 'NOT FOUND':
                        cell = worksheet.cell(row=idx+2, column=file_path_col)
                        cell.hyperlink = file_path
                        cell.style = 'Hyperlink'
            
            # Peace/Love/Hope Orders sheet
            if peace_love_hope_results:
                plh_data = []
                for item in peace_love_hope_results:
                    plh_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'Product': item['product_title'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Variation': item['variation_selected'],
                        'SKU': item['sku'],
                        'Quantity': item['quantity']
                    })
                df_plh = pd.DataFrame(plh_data)
                df_plh.to_excel(writer, sheet_name='Peace-Love-Hope Orders', index=False)
            
            # Other Orders sheet
            if other_orders_results:
                other_data = []
                for item in other_orders_results:
                    other_data.append({
                        'Status': item['order_status'],
                        'Completed': '',
                        'Order ID': item['order_id'],
                        'Customer': item['customer_name'],
                        'Product': item['product_title'],
                        'Price': item['price'],
                        'Message': item['message'],
                        'Variation': item['variation_selected'],
                        'SKU': item['sku'],
                        'Quantity': item['quantity']
                    })
                df_other = pd.DataFrame(other_data)
                df_other.to_excel(writer, sheet_name='Other Orders', index=False)
            
            # All Workflow Orders summary
            if workflow_results['all_orders']:
                df_all = pd.DataFrame(workflow_results['all_orders'])
                df_all.to_excel(writer, sheet_name='All Workflow Orders', index=False)

                # Apply customer grouping colors
                worksheet = writer.sheets['All Workflow Orders']
                self.apply_customer_grouping(worksheet, df_all, 'customer_name')
        
        # Generate Illustrator CSVs
        self.generate_illustrator_csvs(ms_make, ms_update, rr_make, rr_update)
        
        # Create staging folder for slicer batch processing
        if workflow_results['file_locations']:
            self.create_slicer_staging(workflow_results['file_locations'])
        
        self.logger.info(f"Reports generated: {output_file}")
        print(f"\n{'='*60}")
        print(f"Reports saved to: {output_file}")
        print(f"{'='*60}")
        
        # Collect preview count
        preview_count = len([item for item in workflow_results['all_orders'] if item.get('preview_requested', False)])
        if preview_count > 0:
            print(f"Preview Requests: {preview_count}")
            print(f"{'='*60}")
        
        print(f"MS - Needs Made: {len(ms_make)}")
        print(f"MS - Needs Updated: {len(ms_update)}")
        print(f"RR - Needs Made: {len(rr_make)}")
        print(f"RR - Needs Updated: {len(rr_update)}")
        print(f"Peace/Love/Hope Orders: {len(peace_love_hope_results)}")
        print(f"Other Orders: {len(other_orders_results)}")
        print(f"{'='*60}")
    
    def generate_illustrator_csvs(self, ms_make, ms_update, rr_make, rr_update):
        """Generate separate CSVs for MS and RR variants - ONLY items that need to be made from scratch"""

        # MS CSV - only make items (NOT updates)
        if ms_make:
            ms_df = pd.DataFrame([{
                'Status': item['order_status'],
                'Completed': '',
                'Name': item['generated_filename'].split()[0],  # Extract first word (sanitized name without spaces)
                'Center': item['center'],
                'Year': item['year'],
                'Preview': 'preview' if item.get('preview_requested', False) else 'no'
            } for item in ms_make])

            ms_file = self.output_folder / 'illustrator_ms.csv'
            ms_df.to_csv(ms_file, index=False)
            print(f"MS variable CSV: {ms_file.name} ({len(ms_make)} items - Make only)")

        # RR CSV - only make items (NOT updates)
        if rr_make:
            rr_df = pd.DataFrame([{
                'Status': item['order_status'],
                'Completed': '',
                'Name': item['generated_filename'].split()[0],  # Extract first word (sanitized name without spaces)
                'Center': item['center'],
                'Preview': 'preview' if item.get('preview_requested', False) else 'no'
            } for item in rr_make])

            rr_file = self.output_folder / 'illustrator_rr.csv'
            rr_df.to_csv(rr_file, index=False)
            print(f"RR variable CSV: {rr_file.name} ({len(rr_make)} items - Make only)")
        
        # Summary
        total_make = len(ms_make) + len(rr_make)
        total_update = len(ms_update) + len(rr_update)
        ms_work = ms_make + ms_update
        rr_work = rr_make + rr_update
        total_new = len([x for x in (ms_work + rr_work) if x['order_status'] == 'New'])
        print(f"\nSummary:")
        print(f"  Make: {total_make}")
        print(f"  Update: {total_update}")
        print(f"  New Orders: {total_new}")

    def create_slicer_staging(self, file_locations):
        """Create staging folder with files ready for slicer batch processing"""
        import shutil
        import platform
        
        # Determine staging folder based on platform
        system = platform.system()
        if system == "Windows":
            staging_base = Path("C:/Users/tphov/Snowflake_Automation/Slicer_Ready")
        else:  # macOS
            staging_base = Path("/Users/afakename/DocumentsMac/Snowflake_Database/Slicer_Ready")
        
        # Create timestamped batch folder
        batch_folder = staging_base / f"batch_{self.run_timestamp}"
        batch_folder.mkdir(parents=True, exist_ok=True)
        
        print(f"\nCreating slicer staging folder...")
        print("=" * 60)
        
        copied_count = 0
        missing_count = 0
        
        # Copy each file to staging folder
        for item in file_locations:
            file_path = Path(item['file_path'])
            
            if not file_path.exists():
                self.logger.warning(f"Missing file: {file_path}")
                missing_count += 1
                continue
            
            # Copy to staging folder
            destination = batch_folder / file_path.name
            shutil.copy2(file_path, destination)
            copied_count += 1
        
        # Create batch info file
        info_file = batch_folder / "_batch_info.txt"
        with open(info_file, 'w') as f:
            f.write(f"Batch created: {datetime.now()}\n")
            f.write(f"Files ready for slicing: {copied_count}\n")
            f.write(f"Missing files: {missing_count}\n\n")
            f.write("Files in this batch:\n")
            for file in sorted(batch_folder.glob("*.svg")):
                f.write(f"  - {file.name}\n")
        
        print(f"Staging folder: {batch_folder}")
        print(f"Files copied: {copied_count}")
        if missing_count > 0:
            print(f"Missing files: {missing_count}")
        print("=" * 60)    

def main():
    """Main entry point with argument parsing"""
    parser = argparse.ArgumentParser(description='Etsy Automation - Pull and process orders')
    parser.add_argument('--mode', 
                       choices=['time_based', 'open_only'],
                       default='open_only',
                       help='Order fetching mode (default: open_only - status != Complete)')
    parser.add_argument('--days',
                       type=int,
                       default=None,
                       help='Number of days back to fetch (default: 90 for open_only, 30 for time_based)')
    
    args = parser.parse_args()
    
    # Print mode info
    print(f"\n{'='*60}")
    print(f"Etsy Automation Starting")
    print(f"{'='*60}")
    print(f"Mode: {args.mode}")
    if args.mode == 'open_only':
        days = args.days or 90
        print(f"Fetching open orders only (last {days} days, status != Complete)")
    else:
        days = args.days or 30
        print(f"Fetching orders from last {days} days (all statuses)")
    print(f"{'='*60}\n")
    
    # Run automation
    automation = EtsyAutomation(mode=args.mode, days_back=args.days)
    automation.run()

if __name__ == "__main__":
    main()